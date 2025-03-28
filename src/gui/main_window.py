"""
Ventana principal de la aplicación.
Implementa el patrón MVC y Observer.
"""

from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
    QPushButton, QSpinBox, QLabel, QTableWidget, 
    QTableWidgetItem, QHeaderView, QFileDialog, QMessageBox
)
from PyQt6.QtCore import QTimer, Qt
from typing import Dict, Any, Optional
from .widgets.control_panel import PanelControl
from .widgets.process_table import TablaProcesos
from .gantt_widget import DiagramaGantt
from .widgets.metrics_panel import PanelMetricas
from ..core.scheduler import PlanificadorRoundRobin, ObservadorSimulacion
from ..core.process import FabricaProcesos, Proceso
from ..config.settings import (WINDOW_TITLE, WINDOW_MIN_WIDTH, 
                             WINDOW_MIN_HEIGHT, SIMULATION_INTERVAL,
                             SimulationState)
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

class VentanaPrincipal(QMainWindow, ObservadorSimulacion):
    """
    Ventana principal de la aplicación.
    Actúa como controlador en el patrón MVC y como observador de la simulación.
    """
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle(WINDOW_TITLE)
        self.setMinimumSize(WINDOW_MIN_WIDTH, WINDOW_MIN_HEIGHT)
        
        # Inicializar componentes
        self.setup_ui()
        self.setup_timer()
        
        # Estado de la simulación
        self.planificador: Optional[PlanificadorRoundRobin] = None
        self.estado_simulacion = SimulationState.STOPPED
        
        # Datos para el reporte
        self.historial_procesos = []
        self.ultima_metricas = {}
    
    def setup_ui(self):
        """Configura la interfaz de usuario."""
        # Widget central
        widget_central = QWidget()
        self.setCentralWidget(widget_central)
        
        # Layout principal
        layout_principal = QVBoxLayout()
        widget_central.setLayout(layout_principal)
        
        # Panel de control superior
        panel_control = QHBoxLayout()
        layout_principal.addLayout(panel_control)
        
        # Controles de quantum
        label_quantum = QLabel("Quantum:")
        self.spin_quantum = QSpinBox()
        self.spin_quantum.setRange(1, 10)
        self.spin_quantum.setValue(2)
        panel_control.addWidget(label_quantum)
        panel_control.addWidget(self.spin_quantum)
        
        # Controles de procesos
        label_procesos = QLabel("Número de procesos:")
        self.spin_procesos = QSpinBox()
        self.spin_procesos.setRange(1, 10)
        self.spin_procesos.setValue(3)
        panel_control.addWidget(label_procesos)
        panel_control.addWidget(self.spin_procesos)
        
        # Botones
        self.boton_iniciar = QPushButton("Iniciar")
        self.boton_iniciar.clicked.connect(self._iniciar_simulacion)
        self.boton_pausar = QPushButton("Pausar")
        self.boton_pausar.clicked.connect(self._pausar_simulacion)
        self.boton_pausar.setEnabled(False)
        self.boton_exportar = QPushButton("Exportar Reporte")
        self.boton_exportar.clicked.connect(self._exportar_reporte)
        self.boton_exportar.setEnabled(False)
        panel_control.addWidget(self.boton_iniciar)
        panel_control.addWidget(self.boton_pausar)
        panel_control.addWidget(self.boton_exportar)
        
        panel_control.addStretch()
        
        # Tabla de procesos
        self.tabla_procesos = QTableWidget()
        self.tabla_procesos.setColumnCount(7)
        self.tabla_procesos.setHorizontalHeaderLabels([
            "ID", "Llegada", "Duración", "Restante",
            "Estado", "Espera", "Retorno"
        ])
        header = self.tabla_procesos.horizontalHeader()
        for i in range(7):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)
        layout_principal.addWidget(self.tabla_procesos)
        
        # Diagrama de Gantt
        self.diagrama_gantt = DiagramaGantt()
        layout_principal.addWidget(self.diagrama_gantt)
        
        # Panel de métricas
        panel_metricas = QHBoxLayout()
        layout_principal.addLayout(panel_metricas)
        
        self.label_tiempo_total = QLabel("Tiempo total: 0")
        self.label_uso_cpu = QLabel("Uso de CPU: 0%")
        self.label_tiempo_espera = QLabel("T. Espera promedio: 0")
        self.label_tiempo_retorno = QLabel("T. Retorno promedio: 0")
        
        panel_metricas.addWidget(self.label_tiempo_total)
        panel_metricas.addWidget(self.label_uso_cpu)
        panel_metricas.addWidget(self.label_tiempo_espera)
        panel_metricas.addWidget(self.label_tiempo_retorno)
    
    def setup_timer(self):
        """Configura el timer para la simulación."""
        self.timer = QTimer()
        self.timer.timeout.connect(self._tick)
        self.timer.setInterval(SIMULATION_INTERVAL)
    
    def _iniciar_simulacion(self):
        """Inicia una nueva simulación."""
        # Crear planificador
        quantum = self.spin_quantum.value()
        self.planificador = PlanificadorRoundRobin(quantum)
        
        # Crear procesos
        num_procesos = self.spin_procesos.value()
        fabrica = FabricaProcesos()
        for _ in range(num_procesos):
            proceso = fabrica.crear_proceso_aleatorio()
            self.planificador.agregar_proceso(proceso)
        
        # Configurar observadores
        self.planificador.agregar_observador(self)
        self.planificador.agregar_observador(self.diagrama_gantt)
        
        # Limpiar historial
        self.historial_procesos = []
        self.ultima_metricas = {}
        
        # Actualizar UI
        self.boton_iniciar.setEnabled(False)
        self.boton_pausar.setEnabled(True)
        self.boton_exportar.setEnabled(False)
        self.spin_quantum.setEnabled(False)
        self.spin_procesos.setEnabled(False)
        
        # Iniciar timer
        self.timer.start(1000)  # 1 segundo por tick
    
    def _pausar_simulacion(self):
        """Pausa o reanuda la simulación."""
        if self.timer.isActive():
            self.timer.stop()
            self.boton_pausar.setText("Reanudar")
        else:
            self.timer.start()
            self.boton_pausar.setText("Pausar")
    
    def _tick(self):
        """Ejecuta un tick de la simulación."""
        if self.planificador and not self.planificador.tick():
            self.timer.stop()
            self.boton_pausar.setEnabled(False)
            self.boton_iniciar.setEnabled(True)
            self.boton_exportar.setEnabled(True)
            self.spin_quantum.setEnabled(True)
            self.spin_procesos.setEnabled(True)
    
    def actualizar(self, datos: Dict[str, Any]) -> None:
        """
        Actualiza la interfaz con nuevos datos.
        
        Args:
            datos: Diccionario con los datos actualizados
        """
        # Actualizar tabla de procesos
        procesos = datos['procesos']
        self.tabla_procesos.setRowCount(len(procesos))
        
        # Guardar datos para el reporte
        self.historial_procesos = procesos
        self.ultima_metricas = datos.get('metricas', {})
        
        for i, proceso in enumerate(procesos):
            self.tabla_procesos.setItem(i, 0, QTableWidgetItem(str(proceso.id)))
            self.tabla_procesos.setItem(i, 1, QTableWidgetItem(str(proceso.tiempo_llegada)))
            self.tabla_procesos.setItem(i, 2, QTableWidgetItem(str(proceso.tiempo_ejecucion)))
            self.tabla_procesos.setItem(i, 3, QTableWidgetItem(str(proceso.tiempo_restante)))
            self.tabla_procesos.setItem(i, 4, QTableWidgetItem(proceso.estado.name))
            self.tabla_procesos.setItem(i, 5, QTableWidgetItem(
                str(proceso.tiempo_espera) if proceso.tiempo_espera is not None else "-"
            ))
            self.tabla_procesos.setItem(i, 6, QTableWidgetItem(
                str(proceso.tiempo_retorno) if proceso.tiempo_retorno is not None else "-"
            ))
        
        # Actualizar métricas
        metricas = datos['metricas']
        if metricas:
            self.label_tiempo_total.setText(f"Tiempo total: {metricas['tiempo_total']}")
            self.label_uso_cpu.setText(f"Uso de CPU: {metricas['utilizacion_cpu']:.1f}%")
            self.label_tiempo_espera.setText(
                f"T. Espera promedio: {metricas['tiempo_espera_promedio']:.1f}"
            )
            self.label_tiempo_retorno.setText(
                f"T. Retorno promedio: {metricas['tiempo_retorno_promedio']:.1f}"
            )
    
    def _exportar_reporte(self):
        """Exporta los datos de la simulación a un archivo Excel."""
        if not self.historial_procesos:
            QMessageBox.warning(
                self,
                "Advertencia",
                "No hay datos para exportar. Ejecute una simulación primero."
            )
            return
            
        try:
            # Obtener nombre de archivo
            nombre_archivo = f"reporte_simulacion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            ruta_archivo, _ = QFileDialog.getSaveFileName(
                self,
                "Guardar Reporte",
                nombre_archivo,
                "Excel Files (*.xlsx)"
            )
            
            if not ruta_archivo:
                return
            
            # Crear libro de Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte de Simulación"
            
            # Estilos
            titulo_estilo = Font(bold=True, size=12)
            cabecera_estilo = Font(bold=True)
            cabecera_fondo = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            
            # Título
            ws['A1'] = "Reporte de Simulación Round Robin"
            ws['A1'].font = titulo_estilo
            ws.merge_cells('A1:G1')
            ws['A1'].alignment = Alignment(horizontal="center")
            
            # Parámetros de simulación
            ws['A3'] = "Parámetros:"
            ws['A3'].font = cabecera_estilo
            ws['A4'] = f"Quantum: {self.spin_quantum.value() if self.planificador else 'N/A'}"
            ws['A5'] = f"Número de procesos: {len(self.historial_procesos)}"
            
            # Métricas generales
            ws['A7'] = "Métricas Generales:"
            ws['A7'].font = cabecera_estilo
            ws['A8'] = f"Tiempo total: {self.ultima_metricas.get('tiempo_total', 0)}"
            ws['A9'] = f"Utilización CPU: {self.ultima_metricas.get('utilizacion_cpu', 0):.1f}%"
            ws['A10'] = f"Tiempo espera promedio: {self.ultima_metricas.get('tiempo_espera_promedio', 0):.1f}"
            ws['A11'] = f"Tiempo retorno promedio: {self.ultima_metricas.get('tiempo_retorno_promedio', 0):.1f}"
            
            # Tabla de procesos
            ws['A13'] = "Detalle de Procesos:"
            ws['A13'].font = cabecera_estilo
            
            # Cabeceras
            cabeceras = ["ID", "T. Llegada", "T. Ejecución", "T. Restante", 
                        "Estado", "T. Espera", "T. Retorno"]
            for i, cabecera in enumerate(cabeceras, 1):
                celda = ws.cell(row=14, column=i)
                celda.value = cabecera
                celda.font = cabecera_estilo
                celda.fill = cabecera_fondo
            
            # Datos de procesos
            for i, proceso in enumerate(self.historial_procesos, 15):
                ws.cell(row=i, column=1, value=proceso.id)
                ws.cell(row=i, column=2, value=proceso.tiempo_llegada)
                ws.cell(row=i, column=3, value=proceso.tiempo_ejecucion)
                ws.cell(row=i, column=4, value=proceso.tiempo_restante)
                ws.cell(row=i, column=5, value=proceso.estado.name)
                ws.cell(row=i, column=6, value=proceso.tiempo_espera)
                ws.cell(row=i, column=7, value=proceso.tiempo_retorno)
            
            # Ajustar anchos de columna
            for col in range(1, 8):  # A-G
                max_length = 0
                for row in range(1, len(self.historial_procesos) + 16):
                    cell = ws.cell(row=row, column=col)
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = adjusted_width
            
            # Guardar archivo
            wb.save(ruta_archivo)
            
            QMessageBox.information(
                self,
                "Éxito",
                f"Reporte exportado exitosamente a:\n{ruta_archivo}"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"Error al exportar el reporte:\n{str(e)}"
            ) 